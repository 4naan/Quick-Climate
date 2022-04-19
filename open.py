
import datetime
from openpyxl import load_workbook;
#load in excel data
workbook_active_date= load_workbook(filename=r"C:\Users\jcass\Documents\Python\WeatherMe\IDCJDW7007.202204.xlsx");
workbook_historic= load_workbook(filename=r"C:\Users\jcass\Documents\Python\WeatherMe\IDCJCM0033_094008.xlsx");
sheet_active_date = workbook_active_date.active;
sheet_historic = workbook_historic.active;

#enter the date to compare
year = 2022
month = 4 #int(input('Enter a month'))
day = int(input('Enter a day: '))
date_comparison = datetime.datetime(year, month, day)
date_short=datetime.date(year, month, day)
#this finds the dates
selected_row_num=0
end_for_loop=False;
active_dates=[]
for row in sheet_active_date.iter_rows(min_row=8,max_row=26,min_col=2,max_col=2,values_only=True):
    active_dates.append(row)

#find row in dates
for x in active_dates:
    if end_for_loop==False:
        for y in x:
            if y == date_comparison:
                end_for_loop=True;
                break
            selected_row_num=selected_row_num+1

#get the dates highest temp
active_highest=sheet_active_date["D"+str(selected_row_num+8)].value
#get the historic mean temp for month (set to april for the associated dataset)
historic_highest=sheet_historic["E13"].value
if round(active_highest-historic_highest,2)>0:
    print("On the "+str(date_short)+", the highest temperature was +"+str(round(active_highest-historic_highest,2))+" above the historical mean.")
elif round(active_highest-historic_highest,2)<0:
     print("On the "+str(date_short)+", the highest temperature was "+str(round(active_highest-historic_highest,2))+" below the historical mean.")
